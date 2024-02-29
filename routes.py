from flask import Blueprint, request, abort, redirect, url_for, render_template, flash, make_response
from flask_login import login_required, current_user
from werkzeug.security import check_password_hash
from openpyxl import load_workbook, Workbook

import random
import string
import re
import os
import datetime

from data.constans import default_req_state, finished_req

from models import User, Request, Book, License #, Machine, Ownership
from database import db_session
import mail.mail as mail
import data.serverConfig as config

routes = Blueprint('routes', __name__)


def randomPass():
    password = ""
    for i in range(15):
        password += random.choice(string.ascii_letters + string.digits)
    return password


@routes.route('/password', methods=['POST'])
@login_required
def password():
    user = current_user
    password = request.form.get('password')

    if (password != request.form.get('confirmPassword')):
        flash("Las contraseñas no coinciden.")
        return render_template('changePassword.html')
    
    if os.environ.get('ENV') != "development":
        if (check_password_hash(user.password, password)):
            flash("No puedes utilizar la misma contraseña de nuevo.")
            return render_template('changePassword.html')

    try:
        User.setPassword(user,password)
        user.activated = True
        db_session.commit()

        try:
            mail.sendMail(user.email, f"Se ha actualizado tu contraseña en {config.SERVER_NAME}", mail.buildMessage('updatedPassword', {'$ADMIN_MAIL': config.ADMIN_MAIL}))
        except:
            print("Error al enviar el email de confirmación de activación de la cuenta.")
        finally:
            return redirect(url_for('views.index'))

    except:
        return render_template('changePassword.html', error="Ha ocurrido un error.")

#TODO: Habría que ponerlo en un try catch para que la app no muera frente a formularios raros
@routes.route('/solicitar', methods=['POST'])
@login_required
def requests():
    action = request.form['action']
    email = request.form.get('email', current_user.email)
    grouped_data = {}
    for key, value in request.form.items():
        if key in ['action', 'email', 'title', 'message']:
            continue

        field_number = key.split('_')[1]

        if field_number not in grouped_data:
            grouped_data[field_number] = {}

        if key.startswith('isbn'):
            grouped_data[field_number]['isbn'] = value
        elif key.startswith('numlic'):
            grouped_data[field_number]['numlic'] = value

    for req in grouped_data:
        new_req = Request(current_user.id, grouped_data[req]['isbn'], grouped_data[req]['numlic'], default_req_state)
        if current_user.role == "interno" or "promotor" and action == "enviar":
            send_licenses(email, grouped_data[req]['isbn'], grouped_data[req]['numlic'])
        elif current_user.role != "externo" or action != "solicitar":
            abort(403)
        # db_session.add(new_req)
        # db_session.commit()
        
def send_licenses(email, isbn, number):
    import pdb
    pdb.set_trace()
    request.query()
    request.finished()


@routes.route('/licencias', methods=['POST'])
@login_required
def licencias():
    if current_user.admin:
        if request.method == 'POST':
            license_excel = request.files['file']
            workbook = load_workbook(license_excel)
            sheet = workbook.active
            CABECERAS = ('ISBN', 'Libro', 'Código', 'Usuario', 'Fecha caducidad licencias:', 'Duración licencias:')
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                cabeceras_excel = row
            if CABECERAS != cabeceras_excel:
                return redirect(url_for('views.active_licenses', error=True))
            duplicated_licenses = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                isbn, title, license_code, user_type, expiration_date, duration = row
                isbn = isbn.replace("-", "")
                book = Book.query.get(isbn)
                if not book:
                    new_book = Book(isbn, title)
                    db_session.add(new_book)
                    db_session.commit()
                    book = Book.query.get(isbn)
                
                license = License.query.get(license_code)
                if license:
                    duplicated_licenses.append(license)
                else:
                    new_license = License(license_code, isbn, user_type, expiration_date, duration)
                    db_session.add(new_license)
                    db_session.commit()
            # TODO: Hacer esto más bonito
            if duplicated_licenses:
                return 'Página con las licencias duplicadas'
            else:
                return 'Licencias subidas correctamente'

@routes.route('/licencias/descargar', methods=['GET'])
@login_required
def download_licenses():
    if current_user.admin:
        if request.method == 'GET':
            licenses = License.query.all()
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.append(['ISBN', 'Libro', 'Código', 'Usuario', 'Fecha caducidad licencias:', 'Duración licencias:', 'Petición'])
            
            for _, item in enumerate(licenses, start=1):
                # TODO: Añadir el título mediante la relación con el book.
                # TODO: Relación con las peticiones
                worksheet.append([item.code, item.model_book.title, item.isbn, item.user_type, item.expiration_date, item.duration, 'Pedida'])
            excel_file = save_workbook_to_bytes(workbook)

            response = make_response(excel_file)
            now = datetime.datetime.now()
            timestamp = now.strftime("%Y%m%d%H%M%S")
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            response.headers['Content-Disposition'] = 'attachment; filename=licencias'+ timestamp +'.xlsx'
            return response


def save_workbook_to_bytes(workbook):
    """Save workbook to bytes stream."""
    from io import BytesIO
    bytes_stream = BytesIO()
    workbook.save(bytes_stream)
    bytes_stream.seek(0)
    return bytes_stream.getvalue()

# @routes.route('/machines', methods=['POST', 'PUT'])
# @login_required
# def machines():
#     if current_user.admin:
#         if request.method == 'PUT':

#             ipRegex = r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$'
#             macRegex = r'(?:[0-9a-fA-F]){12}'
#             macFullRegex = r'(?:[0-9a-fA-F]:?){12}'
#             portRegex = r'^(0|[1-9][0-9]{0,3}|[1-5][0-9]{4}|6[0-4][0-9]{3}|65[0-4][0-9]{2}|655[0-2][0-9]|6553[0-5])$'

#             name = request.json['name'].strip()
#             mac = request.json['mac'].strip()
#             ip = request.json['ip'].strip()
#             port = request.json['port'].strip()

#             if (re.fullmatch(macRegex, mac)):
#                 mac = list(mac)
#                 mac.insert(10 ,':')
#                 mac.insert(8 ,':')
#                 mac.insert(6 ,':')
#                 mac.insert(4 ,':')
#                 mac.insert(2 ,':')
#                 mac = ''.join(mac)

#             if os.environ.get('ENV') != "development":
#                 if (re.fullmatch(macFullRegex, mac)):
#                     if (re.fullmatch(ipRegex, ip)):
#                         if (re.fullmatch(portRegex, port)):
#                             if Machine.query.filter_by(name = name, mac = mac, ip = ip, port = port).first():
#                                 response = {
#                                     'code': 1,
#                                     'message': "Ya existe una máquina con los mismos datos."
#                                 }

#                             else:
#                                 try:
#                                     new_machine = Machine(name, mac, ip, port)
#                                     db_session.add(new_machine)
#                                     db_session.commit()

#                                     machine = Machine.query.filter_by(name = name, mac = mac, ip = ip, port = port).first()
#                                     response = {
#                                         'code': 0,
#                                         'id': machine.id
#                                     }

#                                 except:
#                                     response = {
#                                         'code': 2,
#                                         'message': "Error al añadir la máquina a la base de datos."
#                                     }

#                         else:
#                             response = {
#                                 'code': 5,
#                                 'message': "El puerto introducido no tiene un formato válido."
#                             }

#                     else:
#                         response = {
#                             'code': 4,
#                             'message': "La dirección IP introducida no tiene un formato válido."
#                         }

#                 else:
#                     response = {
#                         'code': 3,
#                         'message': "La dirección MAC introducida no tiene un formato válido."
#                     }

#             else:
#                 try:
#                     new_machine = Machine(name, mac, ip, port)
#                     db_session.add(new_machine)
#                     db_session.commit()
#                     machine = Machine.query.filter_by(name = name, mac = mac, ip = ip, port = port).first()
#                     response = {
#                         'code': 0,
#                         'id': machine.id
#                     }

#                 except:
#                     response = {
#                         'code': 2,
#                         'message': "Error al añadir la máquina a la base de datos."
#                     }

#             return response
          
#     else:
#         abort(403)


# @routes.route('/machines/<int:machineId>', methods=['POST', 'DELETE'])
# @login_required
# def machine(machineId):
#     if current_user.admin:
#         if request.method == 'POST':
#             update_machine = Machine.query.get(machineId)

#             name = request.json['name']
#             mac = request.json['mac']
#             ip = request.json['ip']
#             port = request.json['port']

#             if Machine.query.filter_by(id=machineId).first():
#                 if not Machine.query.filter_by(name = name, mac = mac, ip = ip, port = port).first():
#                     try:
#                         update_machine.name = name
#                         update_machine.mac = mac
#                         update_machine.ip = ip
#                         update_machine.port = port
#                         db_session.commit()

#                         response = {
#                             'code': 0
#                         }

#                     except:
#                         response = {
#                             'code': 1,
#                             'message': "Error al actualizar la máquina en la base de datos."
#                         }

#                 else:
#                     response = {
#                         'code': 2,
#                         'message': "Ya existe una máquina con los mismos datos."
#                     }
            
#             else:
#                 response = {
#                     'code': 3,
#                     'message': "La máquina especificada no existe."
#                 }

#             return response
        
#         elif request.method == 'DELETE':
#             if Machine.query.filter_by(id=machineId).first():
#                 try:
#                     db_session.delete(Machine.query.get(machineId))
#                     db_session.commit()
#                     response = {
#                             'code': 0
#                     }

#                 except:
#                     response = {
#                         'code': 1,
#                         'message': "Error al eliminar la máquina de la base de datos."
#                     }

#             else:
#                 response = {
#                     'code': 2,
#                     'message': "La máquina especificada no existe."
#                 }
            
#             return response
#     else:
#         abort(403)
        

@routes.route('/users', methods=['PUT'])
@login_required
def users():
    if current_user.admin:

        # Guarda el nuevo usuario especificado en la base de datos.
        if request.method == 'PUT':
            regex = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
            email = request.json['email'].strip()

            if (re.fullmatch(regex, email)):
                if User.query.filter_by(email=request.json['email']).first():
                    response = {
                        'code': 3,
                        'message': "Ya existe un usuario con ese email."
                    }

                else:
                    password = randomPass()
                    newUser = User(email, password)
                    try:
                        db_session.add(newUser)
                        mailResult = mail.sendMail(email, f"Nuevas credenciales en {config.SERVER_NAME}", mail.buildMessage('newUser', {'$PASSWORD': password, '$ADMIN_MAIL': config.ADMIN_MAIL}))
                        if mailResult == 0:
                            db_session.commit()
                            user = User.query.filter_by(email=request.json['email']).first()
                            response = {
                                'code': 0,
                                'id': user.id
                            }
                        else:
                            user = User.query.filter_by(email=request.json['email']).first()
                            if user:
                                db_session.delete(user)
                                db_session.commit()
                            response = {
                            'code': 2,
                            'message': f"Ha ocurrido un error al enviar el correo electrónico. ERROR: {mailResult}"
                        }
                    except:
                        response = {
                            'code': 1,
                            'message': "Error al guardar el usuario en la base de datos."
                        }

            else: 
                response = {
                    'code': 4,
                    'message': "No se ha especificado un email válido."
                }


            return response

    else:
        abort(403)


@routes.route('/users/<int:userId>', methods=['PATCH', 'POST', 'DELETE'])
@login_required
def user(userId):
    if current_user.admin:

        # Resetea la contraseña del usuario especificado.
        if request.method == 'PATCH':
            user = User.query.filter_by(id=userId).first()
            if user:
                try:
                    password = randomPass()
                    User.setPassword(user,password)
                    user.activated = False
                    mailResult = mail.sendMail(user.email, f"Reestablecimiento de contraseña en {config.SERVER_NAME}", mail.buildMessage('resetPassword', {'$PASSWORD': password, '$ADMIN_MAIL': config.ADMIN_MAIL}))
                    if mailResult == 0:
                        db_session.commit()
                        response = {
                            'code': 0
                        }
                    else:
                        response = {
                        'code': 3,
                        'message': f"Ha ocurrido un error al enviar el correo electrónico. ERROR: {mailResult}"
                        }

                except:
                    response = {
                        'code': 2,
                        'message': "Error al resetear la contraseña del usuario."
                    }
            else:
                response = {
                        'code': 1,
                        'message': "El usuario especificado no existe."
                    }

            return response

        # Actualiza los datos del usuario especificado.
        elif request.method == 'POST':
            user = User.query.filter_by(id=userId).first()
            if user:
                try:
                    user.admin = True if request.json['admin'] == 0 else False
                    db_session.commit()
                    isAdmin = '1' if user.admin else '0'
                    response = {
                        'code': 0,
                        'isAdmin': isAdmin
                    }
                except:
                    response = {
                        'code': 2,
                        'message': "Error al cambiar los permisos del usuario."
                    }
            else:
                response = {
                        'code': 1,
                        'message': "El usuario especificado no existe."
                    }

            return response

        
        # Elimina el usuario especificado.
        elif request.method == 'DELETE':
            user = User.query.filter_by(id=userId).first()
            if user:
                try:
                    db_session.delete(user)
                    db_session.commit()
                    response = {
                        'code': 0,
                        'id': userId
                    }
                except:
                    response = {
                        'code': 2,
                        'message': "Error al buscar el usuario en la base de datos."
                    }
            else:
                response = {
                        'code': 1,
                        'message': "El usuario especificado no existe."
                    }

            return response
    else:
        abort(403)


# @routes.route('/users/<int:userId>/machines', methods=['PUT', 'DELETE'])
# @login_required
# def user_machines(userId):
#     if current_user.admin:

#         if request.method == 'PUT':
#             try:
#                 new_ownership = Ownership(user_id=userId, machine_id=request.json["machine_id"])
#                 db_session.add(new_ownership)
#                 db_session.commit()
#                 response = {
#                     'code': 0,
#                     'isOwned': 1
#                 }
#             except:
#                 response = {
#                     'code': 1,
#                     'message': "Error añadiendo propiedad a la máquina {request.json['machine_id']}."
#                 }
#             return response

#         elif request.method == 'DELETE':
#             try:
#                 ownership = Ownership.query.filter_by(user_id = userId, machine_id=request.json["machine_id"]).first()
#                 if ownership:
#                     db_session.delete(ownership)
#                 db_session.commit()
#                 response = {
#                     'code': 0,
#                     'isOwned': 0
#                 }
#             except:
#                 response = {
#                     'code': 1,
#                     'message': f"Error eliminando la propiedad de la máquina {request.json['machine_id']}."
#                 }
#             return response

#     else:
#         abort(403)
#